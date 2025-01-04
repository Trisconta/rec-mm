""" read_recibos.py
(c)2025  Henrique Moreira
"""

import sys
import os.path
import openpyxl
from openpyxl.utils import get_column_letter
sys.path.append(os.path.join(os.path.dirname(__file__), "src", "packages"))
from receipt.config import Config

DEBUG = 1

NOW_YEAR = 2025

MONTHS = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Marco",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

FIELDS_1 = {
    "E7": ("month", "Janeiro", "s"),
    "F7": ("year", 2023, "n"),
    "B9": ("floor", 1, "n"),
    "C9": ("letter", "A", "s"),
    "H9": ("perm", "Permilagem", "s"),
    "I9": ("permil", 0.195, "n"),
    "D14": ("t_value", "Valor do Condominio", "s"),
    "G14": ("condo_v1", 33.5, "n"),
    "D15": ("t_f_res", "Fundo de Reserva", "s"),
    "G15": ("condo_v2", 5.85, "s"),
    "D16": ("t_other", "", "s"),
    "G16": ("condo_v3", 0.0, "n"),
    # Fixed stuff
    "A1": ("designation", "Administracao", "s"),
    "A2": ("nif", "NIF: %s", "s"),
    "A3": ("iban", "PT050xxx", "s"),
    "A4": ("mail", "E-mail: ruamariamachado2@gmail.com", "s"),
}

EAST_2 = {
    "E7": "O7",
    "F7": "P7",
}

SOUTH_2 = {
    "E7": "E34",
    "F7": "F34",
}

def main():
    cfg = Config()
    bdir = cfg.main_path()
    assert bdir, cfg.config_file()
    fname = "../../snippets/xecibos/2023_Recibos_Condominio.xlsx"
    fname = "/opt/local/temp/rec2.xlsx"
    tups = (
        FIELDS_1,
        EAST_2,
        SOUTH_2,
    )
    myhash = build_hash(tups, cfg)
    dump_hash(myhash, "ONE")
    dump_hash(myhash, "TWO")
    read_wbk(fname, myhash)


def coluna(num):
    assert num >= 1, num
    try:
        astr = get_column_letter(num)
    except NameError:
        astr = (chr(ord("A") + num - 1)) if 0 < num <= 26 else "zz"
    return astr

def read_wbk(fname, myhash):
    wbk = openpyxl.open(fname, data_only=True)
    dct = {
        "Workbook": wbk,
        "Sheets": [],
    }
    for s_name in wbk.sheetnames:
        page = wbk[s_name]
        dct["Sheets"].append(page)
    # Check consistency
    is_ok = check_page(myhash, dct["Sheets"][0])
    assert is_ok, "check_page()"
    return wbk, dct

def build_hash(tups, cfg):
    hsh = {}
    hsh = build_west_east(tups, cfg)
    two = build_north_south(tups, cfg)
    fail = False
    myhash = {
        "ONE": hsh,
        "TWO": two,
        "msg": [not fail, "Fail" if fail else "OK"],
    }
    return myhash

def build_west_east(tups, cfg):
    """ Build horizontal, i.e. west-to-east.
    January is FIELDS_1
    """
    fields, east, south = tups
    hsh = {}
    key = sorted(east)[0]
    diff = ord(east[key][0]) - ord(key[0])
    cnt = 1
    for mth in MONTHS:
        hsh[mth] = {}
        for key in fields:
            new = coluna(ord(key[0]) - ord("A") + cnt) + key[1:]
            tup = fields[key]
            fld, data, data_type = tup
            there = data
            if fld == "month":
                data = MONTHS[mth]
            elif fld == "year":
                data = NOW_YEAR
            elif fld == "floor":
                what = 1
                assert data == what, f"month={mth}, {repr(data)} vs {repr(what)}"
            elif fld == "designation":
                data = cfg.get_str("DESIGN")
            elif fld == "nif":
                data = cfg.get_str("NIF")
            elif fld == "iban":
                data = cfg.get_str("IBAN")
            elif fld == "mail":
                what = cfg.get_str("MAIL")
                assert data == what, f"month={mth}, {repr(data)} vs {repr(what)}"
            if mth == 1:
                item = [key, fld, data, data_type]
            else:
                item = [new, fld, data, data_type]
            assert fld not in hsh[mth], f"month={mth}: Dup field: {fld}"
            hsh[mth][fld] = item
        cnt += diff
    return hsh

def build_north_south(tups, cfg):
    """ Build vertical, i.e. north to south.
    """
    fields, east, south = tups
    hsh = {}
    key = sorted(east)[0]
    diff = ord(east[key][0]) - ord(key[0])
    cnt = 1
    for mth in MONTHS:
        hsh[mth] = {}
        for key in fields:
            new = str(int(key[1:]) + cnt)
            tup = fields[key]
            fld, data, data_type = tup
            there = data
            if fld == "month":
                data = MONTHS[mth]
            elif fld == "year":
                data = NOW_YEAR
            elif fld == "floor":
                what = 1
                assert data == what, f"month={mth}, {repr(data)} vs {repr(what)}"
            elif fld == "designation":
                data = cfg.get_str("DESIGN")
            elif fld == "nif":
                data = cfg.get_str("NIF")
            elif fld == "iban":
                data = cfg.get_str("IBAN")
            elif fld == "mail":
                what = cfg.get_str("MAIL")
                assert data == what, f"month={mth}, {repr(data)} vs {repr(what)}"
            if mth == 1:
                item = [key, fld, data, data_type]
            else:
                item = [new, fld, data, data_type]
            assert fld not in hsh[mth], f"month={mth}: Dup field: {fld}"
            hsh[mth][fld] = item
        cnt += diff
    return hsh

def check_page(myhash, page):
    """ Check consistency. """
    hsh = myhash["ONE"]
    fail = False
    for mth in MONTHS:
        print("Checking month:", hsh[mth])
        tup = hsh[mth]["month"]
        cord = tup[0]
        aval = tup[2]
        if mth == 3:
            continue
        assert page[cord].value == aval, f"mth={mth}, {repr(page[cord].value)} != {repr(aval)}"
    tip = {}
    for key in sorted(hsh[1]):
        if key == "month":
            continue
        tip[key] = page[hsh[1][key][0]].value
    for mth in sorted(hsh):
        if mth == 1:
            continue
        # All fields (except month) must be the same to the first (month), i.e. January
        for key in sorted(hsh[mth]):
            value = page[hsh[mth][key][0]].value
            if key not in tip:
                continue
            if tip[key] != value:
                fail = True
                print(
                    ":::", mth, key, not fail, "TIP vs value:",
                    repr(tip[key]), repr(value)
                )
    return not fail

def dump_hash(myhash, hint):
    hsh = myhash[hint]
    if DEBUG <= 0:
        return False
    print(f"dump_hash(): {hint}")
    for key in hsh:
        item = hsh[key]
        for what, cont in item.items():
            print("@" + hint, key, what, cont)
        print()
    return True

main()
